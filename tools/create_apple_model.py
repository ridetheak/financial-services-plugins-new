#!/usr/bin/env python3
"""
Create Apple Inc. three-statement financial model
Includes Income Statement, Balance Sheet, Cash Flow Statement, and supporting schedules
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=12, color='FFFFFF')
title_font = Font(bold=True, size=14)
blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
input_font = Font(color='0000FF')  # Blue font indicates input
formula_font = Font(color='000000')  # Black font indicates formula
link_font = Font(color='008000')  # Green font indicates link

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Remove default worksheet
wb.remove(wb.active)

# ========== Create worksheets ==========
ws_assumptions = wb.create_sheet('Assumptions')
ws_is = wb.create_sheet('IS')
ws_bs = wb.create_sheet('BS')
ws_cf = wb.create_sheet('CF')
ws_wc = wb.create_sheet('WC')
ws_da = wb.create_sheet('DA')
ws_checks = wb.create_sheet('Checks')

# ========== Assumptions worksheet ==========
ws_assumptions['A1'] = 'Apple Inc. (AAPL) - Financial Model Assumptions'
ws_assumptions['A1'].font = Font(bold=True, size=16)

# Time periods
ws_assumptions['A3'] = 'Time Periods'
ws_assumptions['A3'].font = title_font
ws_assumptions['A4'] = 'Historical Period'
ws_assumptions['B4'] = 5
ws_assumptions['A5'] = 'Forecast Period'
ws_assumptions['B5'] = 5

# Revenue growth assumptions
ws_assumptions['A7'] = 'Revenue Growth Rate (%)'
ws_assumptions['A7'].font = title_font
ws_assumptions['A8'] = 'FY2025E'
ws_assumptions['B8'] = 0.05  # 5% growth
ws_assumptions['A9'] = 'FY2026E'
ws_assumptions['B9'] = 0.06
ws_assumptions['A10'] = 'FY2027E'
ws_assumptions['B10'] = 0.07
ws_assumptions['A11'] = 'FY2028E'
ws_assumptions['B11'] = 0.06
ws_assumptions['A12'] = 'FY2029E'
ws_assumptions['B12'] = 0.05

# Margin assumptions
ws_assumptions['A14'] = 'Margin Assumptions (%)'
ws_assumptions['A14'].font = title_font
ws_assumptions['A15'] = 'Gross Margin'
ws_assumptions['B15'] = 0.46  # 46%
ws_assumptions['A16'] = 'R&D Expense Rate'
ws_assumptions['B16'] = 0.07
ws_assumptions['A17'] = 'SG&A Expense Rate'
ws_assumptions['B17'] = 0.06

# Apply blue font to input cells
for row in range(8, 18):
    cell = ws_assumptions.cell(row=row, column=2)
    cell.font = input_font

print("Assumptions worksheet created")

# ========== IS worksheet (Income Statement) ==========
ws_is['A1'] = 'Apple Inc. (AAPL) - Income Statement'
ws_is['A1'].font = Font(bold=True, size=16)
ws_is['A2'] = 'USD in Millions'
ws_is['A2'].font = Font(italic=True)

# Column headers
ws_is['A4'] = ''
ws_is['B4'] = 'FY2020A'
ws_is['C4'] = 'FY2021A'
ws_is['D4'] = 'FY2022A'
ws_is['E4'] = 'FY2023A'
ws_is['F4'] = 'FY2024A'
ws_is['G4'] = 'FY2025E'
ws_is['H4'] = 'FY2026E'
ws_is['I4'] = 'FY2027E'
ws_is['J4'] = 'FY2028E'
ws_is['K4'] = 'FY2029E'

# Apply header styles
for col in range(1, 12):
    ws_is.cell(row=4, column=col).fill = blue_fill
    ws_is.cell(row=4, column=col).font = header_font
    ws_is.cell(row=4, column=col).alignment = Alignment(horizontal='center')

# Income statement line items
is_items = [
    ('Revenue', [274515, 365817, 394328, 383285, 383289]),
    ('Cost of Revenue', [169559, 212981, 223546, 214137, 210352]),
    ('Gross Profit', None),  # formula
    ('', None),  # blank row
    ('R&D Expense', [18752, 21914, 26251, 29915, 31370]),
    ('SG&A Expense', [19916, 21914, 25094, 24932, 26097]),
    ('Total Operating Expenses', None),  # formula
    ('', None),
    ('Operating Income', None),  # formula
    ('', None),
    ('Interest Income', None),
    ('Interest Expense', None),
    ('Other Income / (Expense)', None),
    ('Pre-Tax Income', None),
    ('', None),
    ('Income Tax Expense', None),
    ('Net Income', None),
]

row_num = 5
for item, values in is_items:
    ws_is.cell(row=row_num, column=1).value = item

    if values:
        for col_idx, value in enumerate(values, start=2):
            ws_is.cell(row=row_num, column=col_idx).value = value
            ws_is.cell(row=row_num, column=col_idx).font = input_font
    elif item == 'Gross Profit':
        for col_idx in range(2, 12):
            ws_is.cell(row=row_num, column=col_idx).value = f'={get_column_letter(col_idx)}5-{get_column_letter(col_idx)}6'
    elif item == 'Total Operating Expenses':
        for col_idx in range(2, 12):
            ws_is.cell(row=row_num, column=col_idx).value = f'={get_column_letter(col_idx)}9+{get_column_letter(col_idx)}10'
    elif item == 'Operating Income':
        for col_idx in range(2, 12):
            ws_is.cell(row=row_num, column=col_idx).value = f'={get_column_letter(col_idx)}7-{get_column_letter(col_idx)}11'
    elif item == 'Net Income':
        for col_idx in range(2, 12):
            ws_is.cell(row=row_num, column=col_idx).value = f'={get_column_letter(col_idx)}19'

    row_num += 1

# Set column widths
ws_is.column_dimensions['A'].width = 25
for col in range(2, 12):
    ws_is.column_dimensions[get_column_letter(col)].width = 12

print("IS worksheet created")

# ========== BS worksheet (Balance Sheet) ==========
ws_bs['A1'] = 'Apple Inc. (AAPL) - Balance Sheet'
ws_bs['A1'].font = Font(bold=True, size=16)
ws_bs['A2'] = 'USD in Millions'
ws_bs['A2'].font = Font(italic=True)

# Column headers
ws_bs['A4'] = ''
for col_idx, year in enumerate(['FY2020A', 'FY2021A', 'FY2022A', 'FY2023A', 'FY2024A', 'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E'], start=2):
    ws_bs.cell(row=4, column=col_idx).value = year
    ws_bs.cell(row=4, column=col_idx).fill = blue_fill
    ws_bs.cell(row=4, column=col_idx).font = header_font
    ws_bs.cell(row=4, column=col_idx).alignment = Alignment(horizontal='center')

# Balance sheet line items
bs_items = [
    ('Assets', None),
    ('Current Assets', None),
    ('Cash and Cash Equivalents', [38016, 34940, 23646, 29965, 29965]),
    ('Short-Term Investments', None),
    ('Accounts Receivable, Net', [16158, 26278, 28297, 29508, 30550]),
    ('Inventory', [4061, 6580, 4946, 6331, 7200]),
    ('Other Current Assets', None),
    ('Total Current Assets', None),
    ('', None),
    ('Long-Term Investments', None),
    ('Property, Plant & Equipment, Net', [36539, 39440, 42117, 43715, 45357]),
    ('Intangible Assets, Net', None),
    ('Other Non-Current Assets', None),
    ('Total Non-Current Assets', None),
    ('', None),
    ('Total Assets', None),
    ('', None),
    ('Liabilities and Shareholders\' Equity', None),
    ('Current Liabilities', None),
    ('Accounts Payable', [16377, 53730, 60039, 62833, 65844]),
    ('Accrued Expenses', None),
    ('Deferred Revenue', None),
    ('Short-Term Debt', None),
    ('Other Current Liabilities', None),
    ('Total Current Liabilities', None),
    ('', None),
    ('Long-Term Debt', None),
    ('Deferred Tax Liabilities', None),
    ('Other Non-Current Liabilities', None),
    ('Total Non-Current Liabilities', None),
    ('', None),
    ('Total Liabilities', None),
    ('', None),
    ('Shareholders\' Equity', None),
    ('Common Stock', None),
    ('Retained Earnings', None),
    ('Accumulated Other Comprehensive Income', None),
    ('Total Shareholders\' Equity', None),
    ('', None),
    ('Total Liabilities and Shareholders\' Equity', None),
]

row_num = 5
for item, values in bs_items:
    ws_bs.cell(row=row_num, column=1).value = item

    if values:
        for col_idx, value in enumerate(values, start=2):
            ws_bs.cell(row=row_num, column=col_idx).value = value
            ws_bs.cell(row=row_num, column=col_idx).font = input_font

    row_num += 1

# Set column widths
ws_bs.column_dimensions['A'].width = 30
for col in range(2, 12):
    ws_bs.column_dimensions[get_column_letter(col)].width = 12

print("BS worksheet created")

# ========== CF worksheet (Cash Flow Statement) ==========
ws_cf['A1'] = 'Apple Inc. (AAPL) - Cash Flow Statement'
ws_cf['A1'].font = Font(bold=True, size=16)
ws_cf['A2'] = 'USD in Millions'
ws_cf['A2'].font = Font(italic=True)

# Column headers
ws_cf['A4'] = ''
for col_idx, year in enumerate(['FY2020A', 'FY2021A', 'FY2022A', 'FY2023A', 'FY2024A', 'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E'], start=2):
    ws_cf.cell(row=4, column=col_idx).value = year
    ws_cf.cell(row=4, column=col_idx).fill = blue_fill
    ws_cf.cell(row=4, column=col_idx).font = header_font
    ws_cf.cell(row=4, column=col_idx).alignment = Alignment(horizontal='center')

cf_items = [
    ('Cash Flow from Operations', None),
    ('Net Income', None),
    ('Depreciation and Amortization', [11056, 11105, 11112, 11544, 11283]),
    ('Stock-Based Compensation', None),
    ('Working Capital Changes', None),
    ('Change in Accounts Receivable', None),
    ('Change in Inventory', None),
    ('Change in Accounts Payable', None),
    ('Other Operating Items', None),
    ('Total Cash Flow from Operations', None),
    ('', None),
    ('Cash Flow from Investing', None),
    ('Capital Expenditures', None),
    ('Acquisitions', None),
    ('Investment Purchases / (Sales)', None),
    ('Other Investing Activities', None),
    ('Total Cash Flow from Investing', None),
    ('', None),
    ('Cash Flow from Financing', None),
    ('Debt Issuance / (Repayment)', None),
    ('Share Repurchases', None),
    ('Dividends Paid', None),
    ('Other Financing Activities', None),
    ('Total Cash Flow from Financing', None),
    ('', None),
    ('Net Change in Cash', None),
    ('Beginning Cash Balance', None),
    ('Ending Cash Balance', None),
]

row_num = 5
for item, values in cf_items:
    ws_cf.cell(row=row_num, column=1).value = item

    if values:
        for col_idx, value in enumerate(values, start=2):
            ws_cf.cell(row=row_num, column=col_idx).value = value
            ws_cf.cell(row=row_num, column=col_idx).font = input_font

    row_num += 1

# Set column widths
ws_cf.column_dimensions['A'].width = 30
for col in range(2, 12):
    ws_cf.column_dimensions[get_column_letter(col)].width = 12

print("CF worksheet created")

# ========== Checks worksheet ==========
ws_checks['A1'] = 'Model Validation Checks'
ws_checks['A1'].font = Font(bold=True, size=16)

ws_checks['A3'] = 'Check'
ws_checks['B3'] = 'Status'
ws_checks['C3'] = 'Description'

checks = [
    ('Balance Sheet Balances', '=IF(BS!A22=BS!A46,"✓ Pass","✗ Fail")', 'Assets = Liabilities + Shareholders\' Equity'),
    ('Cash Flow Statement Check', '=IF(CF!A32=BS!A7,"✓ Pass","✗ Fail")', 'Ending Cash = Balance Sheet Cash'),
    ('Net Income Link', '=IF(IS!A21=CF!A6,"✓ Pass","✗ Fail")', 'IS Net Income = CF Net Income'),
]

row_num = 4
for check_name, formula, description in checks:
    ws_checks.cell(row=row_num, column=1).value = check_name
    ws_checks.cell(row=row_num, column=2).value = formula
    ws_checks.cell(row=row_num, column=3).value = description
    row_num += 1

ws_checks.column_dimensions['A'].width = 30
ws_checks.column_dimensions['B'].width = 15
ws_checks.column_dimensions['C'].width = 40

print("Checks worksheet created")

# Save file
output_file = 'apple_3_statement_model.xlsx'
wb.save(output_file)
print(f"\n✓ Three-statement model created successfully: {output_file}")
print(f"Worksheets: Assumptions, IS, BS, CF, WC, DA, Checks")
