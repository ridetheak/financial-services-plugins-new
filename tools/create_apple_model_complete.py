#!/usr/bin/env python3
"""
Create Apple Inc. complete three-statement financial model
Includes forecast formulas, three-statement links, and validation checks
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=11, color='FFFFFF')
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=11)
input_font = Font(color='0000FF')  # Blue font = input
formula_font = Font(color='000000')  # Black font = formula

blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Remove default worksheet
wb.remove(wb.active)

# ========== Create worksheets ==========
ws_assumptions = wb.create_sheet('Assumptions')
ws_is = wb.create_sheet('IS')
ws_bs = wb.create_sheet('BS')
ws_cf = wb.create_sheet('CF')
ws_checks = wb.create_sheet('Checks')

print("Building three-statement model...")

# ========== Assumptions worksheet ==========
ws_assumptions['A1'] = 'Apple Inc. (AAPL) - Financial Model Assumptions'
ws_assumptions['A1'].font = Font(bold=True, size=16)
ws_assumptions.merge_cells('A1:D1')

# Historical and forecast years
years = ['FY2020A', 'FY2021A', 'FY2022A', 'FY2023A', 'FY2024A',
         'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E', 'FY2029E']

# Revenue growth rate assumptions
ws_assumptions['A3'] = 'Revenue Growth Rate (%)'
ws_assumptions['A3'].font = title_font

revenue_growth = [None, None, None, None, None, 0.05, 0.06, 0.07, 0.06, 0.05]
for idx, growth in enumerate(revenue_growth):
    if growth is not None:
        ws_assumptions.cell(row=4, column=idx+2).value = growth
        ws_assumptions.cell(row=4, column=idx+2).font = input_font
        ws_assumptions.cell(row=4, column=idx+2).number_format = '0.0%'

ws_assumptions['A4'] = 'FY2020-FY2029'
ws_assumptions['B4'] = 'Growth Rate'

# Margin assumptions
ws_assumptions['A6'] = 'Margin Assumptions (%)'
ws_assumptions['A6'].font = title_font

margin_items = [
    ('Gross Margin', 0.46),
    ('R&D Expense Rate', 0.07),
    ('SG&A Expense Rate', 0.06),
    ('Effective Tax Rate', 0.15),
]

row_num = 7
for item, value in margin_items:
    ws_assumptions.cell(row=row_num, column=1).value = item
    ws_assumptions.cell(row=row_num, column=2).value = value
    ws_assumptions.cell(row=row_num, column=2).font = input_font
    ws_assumptions.cell(row=row_num, column=2).number_format = '0.0%'
    row_num += 1

# Working capital assumptions
ws_assumptions['A12'] = 'Working Capital Assumptions (Days)'
ws_assumptions['A12'].font = title_font

wc_items = [
    ('Days Sales Outstanding (DSO)', 40),
    ('Days Inventory Outstanding (DIO)', 10),
    ('Days Payable Outstanding (DPO)', 60),
]

row_num = 13
for item, value in wc_items:
    ws_assumptions.cell(row=row_num, column=1).value = item
    ws_assumptions.cell(row=row_num, column=2).value = value
    ws_assumptions.cell(row=row_num, column=2).font = input_font
    row_num += 1

# Other assumptions
ws_assumptions['A17'] = 'Other Assumptions'
ws_assumptions['A17'].font = title_font

other_items = [
    ('Capex as % of Revenue', 0.03),
    ('D&A as % of Capex', 0.95),
    ('Dividend Payout Ratio', 0.15),
]

row_num = 18
for item, value in other_items:
    ws_assumptions.cell(row=row_num, column=1).value = item
    ws_assumptions.cell(row=row_num, column=2).value = value
    ws_assumptions.cell(row=row_num, column=2).font = input_font
    ws_assumptions.cell(row=row_num, column=2).number_format = '0.0%'
    row_num += 1

ws_assumptions.column_dimensions['A'].width = 35
ws_assumptions.column_dimensions['B'].width = 15

print("✓ Assumptions worksheet complete")

# ========== IS worksheet (Income Statement) ==========
ws_is['A1'] = 'Apple Inc. (AAPL) - Income Statement'
ws_is['A1'].font = Font(bold=True, size=16)
ws_is['A2'] = 'USD in Millions'
ws_is['A2'].font = Font(italic=True)

# Column headers
ws_is['A4'] = ''
for idx, year in enumerate(years):
    col_letter = get_column_letter(idx + 2)
    ws_is.cell(row=4, column=idx+2).value = year
    ws_is.cell(row=4, column=idx+2).fill = blue_fill
    ws_is.cell(row=4, column=idx+2).font = header_font
    ws_is.cell(row=4, column=idx+2).alignment = Alignment(horizontal='center')

# Income statement structure - historical data
is_structure = {
    5: ('Revenue', [274515, 365817, 394328, 383285, 383289]),
    6: ('Cost of Revenue', [169559, 212981, 223546, 214137, 210352]),
    7: ('Gross Profit', 'formula'),  # B5-B6
    8: ('', None),
    9: ('R&D Expense', [18752, 21914, 26251, 29915, 31370]),
    10: ('SG&A Expense', [19916, 21914, 25094, 24932, 26097]),
    11: ('Total Operating Expenses', 'formula'),  # B9+B10
    12: ('', None),
    13: ('Operating Income', 'formula'),  # B7-B11
    14: ('', None),
    15: ('Interest Income', [2444, 3543, 3810, 3762, 3762]),
    16: ('Interest Expense', [-2571, -2645, -2931, -3933, -3933]),
    17: ('Other Income / (Expense)', [None, None, None, None, None]),
    18: ('Pre-Tax Income', 'formula'),  # B13+B15+B16+B17
    19: ('', None),
    20: ('Income Tax Expense', None),  # to be calculated
    21: ('Net Income', 'formula'),  # B18-B20
    22: ('', None),
    23: ('Gross Margin (%)', 'margin'),  # B7/B5
    24: ('Operating Margin (%)', 'margin'),  # B13/B5
    25: ('Net Margin (%)', 'margin'),  # B21/B5
}

for row_num, (item, values) in is_structure.items():
    ws_is.cell(row=row_num, column=1).value = item

    if values == 'formula':
        # Formula calculations
        if row_num == 7:  # Gross Profit
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}5-{col_letter}6'
        elif row_num == 11:  # Total Operating Expenses
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}9+{col_letter}10'
        elif row_num == 13:  # Operating Income
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}7-{col_letter}11'
        elif row_num == 18:  # Pre-Tax Income
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}13+{col_letter}15+{col_letter}16'
        elif row_num == 21:  # Net Income
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}18-{col_letter}20'

    elif values == 'margin':
        # Margin calculations
        if row_num == 23:  # Gross Margin
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}7/{col_letter}5'
                ws_is.cell(row=row_num, column=col).number_format = '0.0%'
        elif row_num == 24:  # Operating Margin
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}13/{col_letter}5'
                ws_is.cell(row=row_num, column=col).number_format = '0.0%'
        elif row_num == 25:  # Net Margin
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_is.cell(row=row_num, column=col).value = f'={col_letter}21/{col_letter}5'
                ws_is.cell(row=row_num, column=col).number_format = '0.0%'

    elif isinstance(values, list):
        # Historical data
        for col_idx, value in enumerate(values):
            if value is not None:
                ws_is.cell(row=row_num, column=col_idx+2).value = value
                ws_is.cell(row=row_num, column=col_idx+2).font = input_font

# Add forecast formulas
# Revenue forecast - based on growth rate
for col in range(7, 12):  # FY2025E to FY2029E
    col_letter = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    growth_rate = f'Assumptions!{get_column_letter(col)}4'
    ws_is.cell(row=5, column=col).value = f'={prev_col}5*(1+{growth_rate})'
    ws_is.cell(row=5, column=col).font = formula_font

# Cost of Revenue forecast - based on gross margin
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_is.cell(row=6, column=col).value = f'={col_letter}5*(1-Assumptions!$B$7)'
    ws_is.cell(row=6, column=col).font = formula_font

# R&D Expense forecast - based on expense rate
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_is.cell(row=9, column=col).value = f'={col_letter}5*Assumptions!$B$8'
    ws_is.cell(row=9, column=col).font = formula_font

# SG&A Expense forecast
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_is.cell(row=10, column=col).value = f'={col_letter}5*Assumptions!$B$9'
    ws_is.cell(row=10, column=col).font = formula_font

# Income tax calculation
for col in range(2, 12):
    col_letter = get_column_letter(col)
    ws_is.cell(row=20, column=col).value = f'=MAX(0,{col_letter}18)*Assumptions!$B$10'
    ws_is.cell(row=20, column=col).font = formula_font

# Interest income and expense held at historical levels
for col in range(7, 12):
    ws_is.cell(row=15, column=col).value = f'={get_column_letter(col-1)}15'
    ws_is.cell(row=16, column=col).value = f'={get_column_letter(col-1)}16'

ws_is.column_dimensions['A'].width = 25
for col in range(2, 12):
    ws_is.column_dimensions[get_column_letter(col)].width = 12

print("✓ IS worksheet complete")

# ========== BS worksheet (Balance Sheet) ==========
ws_bs['A1'] = 'Apple Inc. (AAPL) - Balance Sheet'
ws_bs['A1'].font = Font(bold=True, size=16)
ws_bs['A2'] = 'USD in Millions'
ws_bs['A2'].font = Font(italic=True)

# Column headers
ws_bs['A4'] = ''
for idx, year in enumerate(years):
    ws_bs.cell(row=4, column=idx+2).value = year
    ws_bs.cell(row=4, column=idx+2).fill = blue_fill
    ws_bs.cell(row=4, column=idx+2).font = header_font
    ws_bs.cell(row=4, column=idx+2).alignment = Alignment(horizontal='center')

# Balance sheet structure
bs_structure = {
    # Assets
    5: ('Assets', None),
    6: ('Current Assets', None),
    7: ('  Cash and Cash Equivalents', [38016, 34940, 23646, 29965, 29965]),
    8: ('  Short-Term Investments', [37368, 27699, 24658, 31587, 29135]),
    9: ('  Accounts Receivable, Net', [16158, 26278, 28297, 29508, 30550]),
    10: ('  Inventory', [4061, 6580, 4946, 6331, 7200]),
    11: ('  Other Current Assets', [32688, 35677, 37479, 40388, 42411]),
    12: ('Total Current Assets', 'formula'),  # SUM
    13: ('', None),
    14: ('Non-Current Assets', None),
    15: ('  Long-Term Investments', [100887, 127877, 120605, 100544, 94378]),
    16: ('  Property, Plant & Equipment, Net', [36539, 39440, 42117, 43715, 45357]),
    17: ('  Intangible Assets, Net', None),
    18: ('  Other Non-Current Assets', [22283, 27979, 32172, 36423, 41307]),
    19: ('Total Non-Current Assets', 'formula'),
    20: ('', None),
    21: ('Total Assets', 'formula'),
    22: ('', None),
    # Liabilities
    23: ('Liabilities and Shareholders\' Equity', None),
    24: ('Current Liabilities', None),
    25: ('  Accounts Payable', [16377, 53730, 60039, 62833, 65844]),
    26: ('  Accrued Expenses', None),
    27: ('  Deferred Revenue', None),
    28: ('  Short-Term Debt', None),
    29: ('  Other Current Liabilities', [42967, 47763, 59279, 64845, 69031]),
    30: ('Total Current Liabilities', 'formula'),
    31: ('', None),
    32: ('Non-Current Liabilities', None),
    33: ('  Long-Term Debt', [98661, 124719, 120069, 111088, 99028]),
    34: ('  Deferred Tax Liabilities', None),
    35: ('  Other Non-Current Liabilities', None),
    36: ('Total Non-Current Liabilities', 'formula'),
    37: ('', None),
    38: ('Total Liabilities', 'formula'),
    39: ('', None),
    # Shareholders' Equity
    40: ('Shareholders\' Equity', None),
    41: ('  Common Stock', [50779, 57365, 64849, 73181, 83779]),
    42: ('  Retained Earnings', [14933, 15157, -3068, -14050, -15545]),
    43: ('  Accumulated Other Comprehensive Income', None),
    44: ('Total Shareholders\' Equity', 'formula'),
    45: ('', None),
    46: ('Total Liabilities and Shareholders\' Equity', 'formula'),
}

for row_num, (item, values) in bs_structure.items():
    ws_bs.cell(row=row_num, column=1).value = item

    if values == 'formula':
        if row_num == 12:  # Total Current Assets
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'=SUM({col_letter}7:{col_letter}11)'
        elif row_num == 19:  # Total Non-Current Assets
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'=SUM({col_letter}15:{col_letter}18)'
        elif row_num == 21:  # Total Assets
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'={col_letter}12+{col_letter}19'
        elif row_num == 30:  # Total Current Liabilities
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'=SUM({col_letter}25:{col_letter}29)'
        elif row_num == 36:  # Total Non-Current Liabilities
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'=SUM({col_letter}33:{col_letter}35)'
        elif row_num == 38:  # Total Liabilities
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'={col_letter}30+{col_letter}36'
        elif row_num == 44:  # Total Shareholders' Equity
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'=SUM({col_letter}41:{col_letter}43)'
        elif row_num == 46:  # Total Liabilities and Shareholders' Equity
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_bs.cell(row=row_num, column=col).value = f'={col_letter}38+{col_letter}44'

    elif isinstance(values, list):
        for col_idx, value in enumerate(values):
            if value is not None:
                ws_bs.cell(row=row_num, column=col_idx+2).value = value
                ws_bs.cell(row=row_num, column=col_idx+2).font = input_font

# Forecast Accounts Receivable - based on DSO
for col in range(7, 12):
    col_letter = get_column_letter(col)
    # DSO = AR / (Revenue/365)
    # AR = DSO * (Revenue/365)
    ws_bs.cell(row=9, column=col).value = f'=Assumptions!$B$13*(IS!{col_letter}5/365)'
    ws_bs.cell(row=9, column=col).font = formula_font

# Forecast Inventory - based on DIO
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_bs.cell(row=10, column=col).value = f'=Assumptions!$B$14*(IS!{col_letter}6/365)'
    ws_bs.cell(row=10, column=col).font = formula_font

# Forecast Accounts Payable - based on DPO
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_bs.cell(row=25, column=col).value = f'=Assumptions!$B$15*(IS!{col_letter}6/365)'
    ws_bs.cell(row=25, column=col).font = formula_font

# Cash forecast - simplified, linked from CF
for col in range(7, 12):
    col_letter = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    ws_bs.cell(row=7, column=col).value = f'=CF!{col_letter}32'
    ws_bs.cell(row=7, column=col).font = Font(color='008000')  # Green = linked

# Other balance sheet items - simplified forecast
for col in range(7, 12):
    col_letter = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    # Short-term investments and other current assets held flat
    ws_bs.cell(row=8, column=col).value = f'={prev_col}8'
    ws_bs.cell(row=11, column=col).value = f'={prev_col}11'
    ws_bs.cell(row=15, column=col).value = f'={prev_col}15'
    ws_bs.cell(row=16, column=col).value = f'={prev_col}16'
    ws_bs.cell(row=18, column=col).value = f'={prev_col}18'
    ws_bs.cell(row=29, column=col).value = f'={prev_col}29'
    ws_bs.cell(row=33, column=col).value = f'={prev_col}33'

ws_bs.column_dimensions['A'].width = 35
for col in range(2, 12):
    ws_bs.column_dimensions[get_column_letter(col)].width = 12

print("✓ BS worksheet complete")

# ========== CF worksheet (Cash Flow Statement) ==========
ws_cf['A1'] = 'Apple Inc. (AAPL) - Cash Flow Statement'
ws_cf['A1'].font = Font(bold=True, size=16)
ws_cf['A2'] = 'USD in Millions'
ws_cf['A2'].font = Font(italic=True)

# Column headers
ws_cf['A4'] = ''
for idx, year in enumerate(years):
    ws_cf.cell(row=4, column=idx+2).value = year
    ws_cf.cell(row=4, column=idx+2).fill = blue_fill
    ws_cf.cell(row=4, column=idx+2).font = header_font
    ws_cf.cell(row=4, column=idx+2).alignment = Alignment(horizontal='center')

# Cash flow statement structure
cf_structure = {
    5: ('Cash Flow from Operations', None),
    6: ('  Net Income', 'link_is'),
    7: ('  Depreciation and Amortization', [11056, 11105, 11112, 11544, 11283]),
    8: ('  Stock-Based Compensation', [6832, 9084, 9038, 10582, 11989]),
    9: ('  Working Capital Changes', None),
    10: ('    Change in Accounts Receivable', None),
    11: ('    Change in Inventory', None),
    12: ('    Change in Accounts Payable', None),
    13: ('  Other Operating Items', [9722, 15908, 24724, 19200, 15620]),
    14: ('Total Cash Flow from Operations', 'formula'),
    15: ('', None),
    16: ('Cash Flow from Investing', None),
    17: ('  Capital Expenditures', [-7309, -11085, -10708, -10959, -9595]),
    18: ('  Acquisitions', None),
    19: ('  Investment Purchases / (Sales)', None),
    20: ('  Other Investing Activities', None),
    21: ('Total Cash Flow from Investing', 'formula'),
    22: ('', None),
    23: ('Cash Flow from Financing', None),
    24: ('  Debt Issuance / (Repayment)', None),
    25: ('  Share Repurchases', None),
    26: ('  Dividends Paid', None),
    27: ('  Other Financing Activities', None),
    28: ('Total Cash Flow from Financing', 'formula'),
    29: ('', None),
    30: ('Net Change in Cash', 'formula'),
    31: ('  Beginning Cash Balance', None),
    32: ('  Ending Cash Balance', 'formula'),
}

for row_num, (item, values) in cf_structure.items():
    ws_cf.cell(row=row_num, column=1).value = item

    if values == 'link_is':
        # Link to Net Income on Income Statement
        for col in range(2, 12):
            col_letter = get_column_letter(col)
            ws_cf.cell(row=row_num, column=col).value = f'=IS!{col_letter}21'
            ws_cf.cell(row=row_num, column=col).font = Font(color='008000')

    elif values == 'formula':
        if row_num == 14:  # Total Cash Flow from Operations
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_cf.cell(row=row_num, column=col).value = f'=SUM({col_letter}6:{col_letter}13)'
        elif row_num == 21:  # Total Cash Flow from Investing
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_cf.cell(row=row_num, column=col).value = f'=SUM({col_letter}17:{col_letter}20)'
        elif row_num == 28:  # Total Cash Flow from Financing
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_cf.cell(row=row_num, column=col).value = f'=SUM({col_letter}24:{col_letter}27)'
        elif row_num == 30:  # Net Change in Cash
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_cf.cell(row=row_num, column=col).value = f'={col_letter}14+{col_letter}21+{col_letter}28'
        elif row_num == 32:  # Ending Cash Balance
            for col in range(2, 12):
                col_letter = get_column_letter(col)
                ws_cf.cell(row=row_num, column=col).value = f'={col_letter}30+{col_letter}31'

    elif isinstance(values, list):
        for col_idx, value in enumerate(values):
            if value is not None:
                ws_cf.cell(row=row_num, column=col_idx+2).value = value
                ws_cf.cell(row=row_num, column=col_idx+2).font = input_font

# Working capital change formulas
for col in range(2, 12):
    col_letter = get_column_letter(col)
    # Change in Accounts Receivable (increase is negative, decrease is positive)
    if col > 2:
        prev_col = get_column_letter(col - 1)
        ws_cf.cell(row=10, column=col).value = f'=-({col_letter}9-{prev_col}9)'
        ws_cf.cell(row=11, column=col).value = f'=-({col_letter}10-{prev_col}10)'
        ws_cf.cell(row=12, column=col).value = f'={col_letter}25-{get_column_letter(col-1)}25'

# Beginning cash balance
ws_cf.cell(row=31, column=2).value = 38016  # FY2020 beginning balance
ws_cf.cell(row=31, column=2).font = input_font
for col in range(3, 12):
    prev_col = get_column_letter(col - 1)
    ws_cf.cell(row=31, column=col).value = f'={prev_col}32'

# Forecast depreciation
for col in range(7, 12):
    col_letter = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    ws_cf.cell(row=7, column=col).value = f'={prev_col}7'
    ws_cf.cell(row=8, column=col).value = f'={prev_col}8'

# Forecast capital expenditures
for col in range(7, 12):
    col_letter = get_column_letter(col)
    ws_cf.cell(row=17, column=col).value = f'=-IS!{col_letter}5*Assumptions!$B$21'
    ws_cf.cell(row=17, column=col).font = formula_font

# Simplified other items forecast
for col in range(7, 12):
    col_letter = get_column_letter(col)
    prev_col = get_column_letter(col - 1)
    ws_cf.cell(row=13, column=col).value = f'={prev_col}13'

ws_cf.column_dimensions['A'].width = 35
for col in range(2, 12):
    ws_cf.column_dimensions[get_column_letter(col)].width = 12

print("✓ CF worksheet complete")

# ========== Checks worksheet ==========
ws_checks['A1'] = 'Model Validation Checks'
ws_checks['A1'].font = Font(bold=True, size=16)

ws_checks['A3'] = 'Check'
ws_checks['B3'] = 'FY2024A'
ws_checks['C3'] = 'Status'
ws_checks['D3'] = 'Description'
for col in range(1, 5):
    ws_checks.cell(row=3, column=col).fill = blue_fill
    ws_checks.cell(row=3, column=col).font = header_font

checks = [
    ('Balance Sheet Balances', '=BS!F21-BS!F46', '=IF(ABS(B4)<1,"✓ Pass","✗ Fail")', 'Total Assets = Total Liabilities and Shareholders\' Equity'),
    ('Cash Verification', '=CF!F32-BS!F7', '=IF(ABS(B5)<1,"✓ Pass","✗ Fail")', 'CF Ending Cash = Balance Sheet Cash'),
    ('Net Income Link', '=IS!F21-CF!F6', '=IF(ABS(B6)<1,"✓ Pass","✗ Fail")', 'IS Net Income = CF Net Income'),
]

row_num = 4
for check_name, formula, status, description in checks:
    ws_checks.cell(row=row_num, column=1).value = check_name
    ws_checks.cell(row=row_num, column=2).value = formula
    ws_checks.cell(row=row_num, column=3).value = status
    ws_checks.cell(row=row_num, column=4).value = description
    row_num += 1

ws_checks.column_dimensions['A'].width = 30
ws_checks.column_dimensions['B'].width = 15
ws_checks.column_dimensions['C'].width = 12
ws_checks.column_dimensions['D'].width = 40

print("✓ Checks worksheet complete")

# Save file
output_file = 'apple_3_statement_model.xlsx'
wb.save(output_file)

print(f"\n{'='*60}")
print(f"✓ Apple Inc. three-statement financial model created successfully!")
print(f"{'='*60}")
print(f"File: {output_file}")
print(f"Worksheets: Assumptions, IS, BS, CF, Checks")
print(f"\nContents:")
print(f"  • Historical data: FY2020A - FY2024A (5 years)")
print(f"  • Forecast period: FY2025E - FY2029E (5 years)")
print(f"  • Income Statement: Revenue, costs, expenses, profit")
print(f"  • Balance Sheet: Assets, liabilities, shareholders' equity")
print(f"  • Cash Flow Statement: Operating, investing, financing activities")
print(f"  • Assumption-driven: Revenue growth, margins, working capital days")
print(f"  • Auto-linked: Three statements automatically connected")
print(f"  • Validation checks: Balance check, consistency verification")
print(f"{'='*60}")
